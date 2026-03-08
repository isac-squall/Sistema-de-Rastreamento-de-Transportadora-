"""
Módulo para processamento e atualização de planilhas Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from config import (
    EXCEL_FILE_PATH,
    COLUNA_NF,
    COLUNA_RASTREAMENTO,
    COLUNA_STATUS,
    COLUNA_ULTIMA_ATUALIZACAO,
    COLUNA_DETALHES,
    ATUALIZAR_APENAS_MUDANCAS
)
from logger_setup import logger


class ProcessadorExcel:
    """Gerencia leitura e escrita em planilhas Excel"""

    CORES_STATUS = {
        '✅': 'C6EFCE',  # Verde - Entregue
        '🚚': 'FFC7CE',  # Vermelho claro - Em Trânsito
        '📦': 'FFEB9C',  # Amarelo - Saiu para Entrega
        '⏳': 'BDD7EE',  # Azul - Aguardando
        '❌': 'F4B084',  # Laranja - Não Entregue
        '↩️': 'E2EFDA',   # Verde claro - Devolvido
        '🚫': 'D9D9D9',  # Cinza - Cancelado
        '⚠️': 'F8CBAD'    # Laranja claro - Erro
    }

    def __init__(self, caminho_arquivo: str = EXCEL_FILE_PATH):
        """
        Inicializa o processador de Excel

        Args:
            caminho_arquivo: Caminho do arquivo Excel
        """
        self.caminho_arquivo = caminho_arquivo
        self.df = None
        self.wb = None
        self.ws = None

    def carregar_planilha(self) -> bool:
        """
        Carrega a planilha Excel

        Returns:
            True se carregou com sucesso, False caso contrário
        """
        try:
            # Carrega com pandas
            self.df = pd.read_excel(self.caminho_arquivo, dtype=str)
            logger.info(f"Planilha carregada com sucesso: {self.caminho_arquivo}")
            logger.info(f"Linhas: {len(self.df)}, Colunas: {len(self.df.columns)}")

            # Mostra colunas disponíveis
            logger.debug(f"Colunas disponíveis: {list(self.df.columns)}")

            return True

        except FileNotFoundError:
            logger.error(f"Arquivo não encontrado: {self.caminho_arquivo}")
            return False
        except Exception as e:
            logger.error(f"Erro ao carregar planilha: {str(e)}")
            return False

    def validar_colunas_obrigatorias(self) -> bool:
        """
        Valida se as colunas obrigatórias existem

        Returns:
            True se todas as colunas existem, False caso contrário
        """
        colunas_obrigatorias = [COLUNA_NF, COLUNA_RASTREAMENTO, COLUNA_STATUS]
        colunas_faltantes = [col for col in colunas_obrigatorias if col not in self.df.columns]

        if colunas_faltantes:
            logger.error(f"Colunas obrigatórias faltando: {colunas_faltantes}")
            logger.info(f"Colunas esperadas: {colunas_obrigatorias}")
            logger.info(f"Colunas disponíveis: {list(self.df.columns)}")
            return False

        return True

    def obter_registros_para_processar(self) -> List[Dict]:
        """
        Obtém lista de registros para processar

        Returns:
            Lista de dicionários com dados dos registros
        """
        registros = []

        for idx, row in self.df.iterrows():
            nf = str(row.get(COLUNA_NF, '')).strip()
            rastreamento = str(row.get(COLUNA_RASTREAMENTO, '')).strip()

            # Pula linhas sem NF e sem rastreamento
            if not nf and not rastreamento:
                continue

            registros.append({
                'indice': idx,
                'numero_nf': nf if nf and nf != 'nan' else '',
                'codigo_rastreamento': rastreamento if rastreamento and rastreamento != 'nan' else '',
                'status_atual': str(row.get(COLUNA_STATUS, '')).strip(),
                'ultima_atualizacao': str(row.get(COLUNA_ULTIMA_ATUALIZACAO, '')).strip()
            })

        logger.info(f"Total de registros para processar: {len(registros)}")
        return registros

    def atualizar_registro(self, indice_linha: int, dados_rastreamento: Dict) -> bool:
        """
        Atualiza um registro na planilha com dados de rastreamento

        Args:
            indice_linha: Índice da linha a atualizar
            dados_rastreamento: Dicionário com dados de rastreamento

        Returns:
            True se houve mudança de status, False caso contrário
        """
        if dados_rastreamento is None:
            nova_status = '⚠️ Erro na Busca'
        else:
            nova_status = dados_rastreamento.get('status_amigavel', '⚠️ Erro na Busca')

        status_anterior = str(self.df.iloc[indice_linha].get(COLUNA_STATUS, '')).strip()

        # Se foi configurado para atualizar apenas mudanças e não houve mudança, pula
        if ATUALIZAR_APENAS_MUDANCAS and status_anterior == nova_status:
            return False

        # Atualiza status
        self.df.at[indice_linha, COLUNA_STATUS] = nova_status

        # Atualiza última atualização
        self.df.at[indice_linha, COLUNA_ULTIMA_ATUALIZACAO] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Atualiza detalhes se coluna existir
        if COLUNA_DETALHES in self.df.columns and dados_rastreamento:
            detalhes = self._montar_detalhes(dados_rastreamento)
            self.df.at[indice_linha, COLUNA_DETALHES] = detalhes

        return True  # Houve mudança

    def _montar_detalhes(self, dados_rastreamento: Dict) -> str:
        """
        Monta string de detalhes do rastreamento

        Args:
            dados_rastreamento: Dicionário com dados de rastreamento

        Returns:
            String com detalhes formatados
        """
        detalhes_lista = []

        if dados_rastreamento.get('localizacao'):
            detalhes_lista.append(f"📍 {dados_rastreamento['localizacao']}")

        if dados_rastreamento.get('data_atualizacao'):
            detalhes_lista.append(f"🕐 {dados_rastreamento['data_atualizacao']}")

        if dados_rastreamento.get('detalhes', {}).get('endereco_entrega'):
            detalhes_lista.append(f"🏠 {dados_rastreamento['detalhes']['endereco_entrega']}")

        return ' | '.join(detalhes_lista) if detalhes_lista else ''

    def salvar_planilha(self, aplicar_formatacao: bool = True) -> bool:
        """
        Salva as alterações na planilha

        Args:
            aplicar_formatacao: Se deve aplicar formatação visual

        Returns:
            True se salvou com sucesso, False caso contrário
        """
        try:
            # Salva com pandas
            self.df.to_excel(self.caminho_arquivo, index=False)
            logger.info(f"Planilha salva com sucesso: {self.caminho_arquivo}")

            # Aplica formatação se solicitado
            if aplicar_formatacao:
                self._aplicar_formatacao()

            return True

        except Exception as e:
            logger.error(f"Erro ao salvar planilha: {str(e)}")
            return False

    def _aplicar_formatacao(self):
        """Aplica formatação visual à planilha"""
        try:
            wb = load_workbook(self.caminho_arquivo)
            ws = wb.active

            # Formata colunas
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    # Aplica cor baseada no status
                    if cell.column_letter == self._obter_coluna_status():
                        for emoji, cor in self.CORES_STATUS.items():
                            if emoji in str(cell.value or ''):
                                cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
                                cell.font = Font(bold=True)
                                break

            # Ajusta largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value or ''))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.caminho_arquivo)
            logger.debug("Formatação aplicada com sucesso")

        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação: {str(e)}")

    def _obter_coluna_status(self) -> str:
        """Obtém a letra da coluna de status"""
        try:
            indice_coluna = list(self.df.columns).index(COLUNA_STATUS)
            return chr(65 + indice_coluna)  # A=65 na tabela ASCII
        except:
            return 'C'  # Padrão

    def criar_planilha_modelo(self) -> bool:
        """
        Cria uma planilha modelo com estrutura básica

        Returns:
            True se criou com sucesso, False caso contrário
        """
        try:
            df_modelo = pd.DataFrame({
                COLUNA_NF: ['', '', ''],
                COLUNA_RASTREAMENTO: ['', '', ''],
                COLUNA_STATUS: ['⏳ Aguardando Retirada', '🚚 Em Trânsito', '✅ Entregue'],
                COLUNA_ULTIMA_ATUALIZACAO: [
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                ],
                COLUNA_DETALHES: ['', '', '']
            })

            df_modelo.to_excel(self.caminho_arquivo, index=False)
            logger.info(f"Planilha modelo criada: {self.caminho_arquivo}")

            self._aplicar_formatacao()

            return True

        except Exception as e:
            logger.error(f"Erro ao criar planilha modelo: {str(e)}")
            return False

    def gerar_relatorio_atualizacoes(self, atualizacoes: Dict) -> str:
        """
        Gera relatório das atualizações realizadas

        Args:
            atualizacoes: Dicionário com estatísticas de atualizações

        Returns:
            String com relatório formatado
        """
        relatorio = f"""
{'='*80}
                    RELATÓRIO DE ATUALIZAÇÃO DE RASTREAMENTOS
{'='*80}

Data/Hora do Processamento: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

ESTATÍSTICAS:
  ✓ Total de registros processados: {atualizacoes.get('total_processados', 0)}
  ✓ Registros com sucesso: {atualizacoes.get('sucesso', 0)}
  ✓ Registros com mudança de status: {atualizacoes.get('mudancas', 0)}
  ✓ Registros com erro: {atualizacoes.get('erros', 0)}
  
DISTRIBUIÇÃO DE STATUS:
"""
        if atualizacoes.get('status_distribuicao'):
            for status, quantidade in atualizacoes['status_distribuicao'].items():
                relatorio += f"  • {status}: {quantidade}\n"

        relatorio += f"""
DETALHES DE ERROS:
  - Total: {len(atualizacoes.get('erros_detalhes', []))}
"""
        if atualizacoes.get('erros_detalhes'):
            for erro in atualizacoes['erros_detalhes'][:5]:  # Mostra apenas os 5 primeiros
                relatorio += f"  • {erro}\n"
            if len(atualizacoes['erros_detalhes']) > 5:
                relatorio += f"  • ... e mais {len(atualizacoes['erros_detalhes']) - 5} erro(s)\n"

        relatorio += f"""
{'='*80}
Arquivo salvo em: {self.caminho_arquivo}
{'='*80}
"""
        return relatorio
